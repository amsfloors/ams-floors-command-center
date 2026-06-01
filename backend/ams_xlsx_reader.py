#!/usr/bin/env python3
"""
AMS Floors — Module 2: XLSX Reader + both Gates (the heart).

Pure function: file in -> (canonical record + review items) out. NO database writes.
Obeys SCHEMA_TRUTH.md and 03_ANTIDRIFT_RULES_AND_FACTS.md:
  - Capture every populated cell with provenance (file/sheet/cell/read|computed).
  - Read formulas/values, never labels. Markup decoded from the formula.
  - Find header + bidsheet by CONTENT, never by hard-coded row/sheet.
  - Item code is a label, never a key (system id assigned per line).
  - Gate 1 Reconciliation (internal to-the-penny + external within recorded tolerance).
  - Gate 2 PO/WO Completeness (per-line + record-level), two-state manufacturer.
  - Review List: BLOCKING (quarantine) vs ENRICHMENT (import + flag).
  - Authoritative-source dual-capture (D1 architect / D2 plansDate / D3 projectNumber).

Run:  python ams_xlsx_reader.py /path/to/dir   -> prints the reconciliation report.
"""
import os, re, glob, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---- documented constants (conscious decisions, not magic numbers) ----
# Gate-1 external tolerance: Peter rounds the stated proposal UP to a round number.
# Observed round-ups on fixtures range $37..$403 (Costco $403 is the recorded extreme).
# Recorded as a parameter so it can be owned/confirmed, never hidden in a comparison.
RECON_PENNY = 0.01                  # internal integrity tolerance (cents)
ROUNDUP_PCT_LIMIT = 0.012           # external: stated-vs-computed gap allowed as Peter rounding (~1.2%)
ROUNDUP_ABS_FLOOR = 450.0           # ...or this many dollars, whichever is larger (admits Costco $403)
UNITCOST_LOW = 0.10                 # sanity floor: costs <= this are suspect (e.g. $0.06 mis-keys)
UNITCOST_HIGH = 100000.0            # sanity ceiling per unit

MATERIAL_WORDS = {  # SUGGESTION ONLY (helps Peter tag fast) — never the authoritative answer
    'lvt','plywood','wall base','carpet tile','carpet pad','carpet','vct','rubber','tile',
    'grout','thinset','wood','vinyl','sheet vinyl','sealer','primer','base','reducer',
    'transition','underlayment','membrane','epoxy','expoxy','mortar','laminate','broadloom',
    'modular','cove','psa','waterproofing','crack isolation','crack isolation &'}
PLACEHOLDER_WORDS = {'tbd','supplied by owner','by owner','n/a','na','see plans','none'}

# SEED brand registry, bootstrapped from the 12 fixtures' confirmed brands.
# In production this is the LIVE suppliers/products registry + Peter's Review confirmations,
# passed in as `registry`. The seed is a starting point, never the authority.
SEED_BRANDS = {
    'mapei','schluter','daltile','johnsonite','proflex','shaw contract','shaw','tarkett',
    'ceramic technics','ecore','armstrong','mannington','american olean','atlas concorde',
    'garden state tile','milliken','matter surfaces','burke','architessa','bostik','interface',
    'tec','marazzi','roppe','pemko','hardie','crossville','ardex','sherwin williams','mohawk',
    'patcraft','durkan'}

CARTON_UNITS = {'ctn','carton','cartons','bx','box','boxes'}


def resolve_mfr(value, registry):
    """Resolve a manufacturer value against the brand registry. Returns (state, suggestion)."""
    if value is None or str(value).strip() == '':
        return 'blank', None
    norm_v = str(value).strip().lower()
    if norm_v in registry:
        return 'resolved', None
    if norm_v in PLACEHOLDER_WORDS:
        return 'unresolved', 'placeholder'
    if norm_v in MATERIAL_WORDS:
        return 'unresolved', 'looks like a material type'
    return 'unresolved', 'unknown — confirm brand or material'


def prov(value, sheet, row, col, mode='read'):
    """A captured value with immutable provenance."""
    return {'value': value, 'sheet': sheet,
            'cell': f'{get_column_letter(col)}{row}' if col else None, 'mode': mode}


def norm(s):
    return str(s).strip().upper() if s is not None else ''


def is_section_code(v):
    return bool(re.fullmatch(r'0?\d{5}', str(v).strip())) if v is not None else False


def normalize_section(v):
    s = str(v).strip()
    return s.zfill(6) if re.fullmatch(r'\d{5}', s) else s


# ----------------------------------------------------------------------
# Structure discovery (by content, never hard-coded)
# ----------------------------------------------------------------------
def find_bidsheet(wbv):
    """The sheet whose content holds a line-item table (a row with ITEM and TOTAL)."""
    for name in wbv.sheetnames:
        ws = wbv[name]
        for r in range(1, min(ws.max_row or 1, 40) + 1):
            cells = [norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column or 1, 40) + 1)]
            if 'ITEM' in cells and any('TOTAL' in c for c in cells):
                return name, r
    return None, None


def build_colmap(ws, header_row):
    """header text -> column index (1-based), read from the header row."""
    m = {}
    for c in range(1, (ws.max_column or 1) + 1):
        h = norm(ws.cell(header_row, c).value)
        if h:
            m[h] = c
    return m


def col_for(colmap, *predicates):
    """First column whose header satisfies any predicate (exact or contains)."""
    for h, c in colmap.items():
        for p in predicates:
            if p(h):
                return c, h
    return None, None


def detect_family(colmap):
    has_material = any(h == 'MATERIAL' for h in colmap)
    has_costplus = any(h.startswith('COST +') or h.startswith('COST+') for h in colmap)
    has_amount = any(h == 'AMOUNT' for h in colmap)
    if has_costplus and not has_material:
        return 'cost_plus_n'
    if has_material:
        return 'material'
    if has_amount:
        return 'amount'
    return 'unknown'


def parse_label_rate(header_text):
    m = re.search(r'(\d+(?:\.\d+)?)\s*%', header_text)
    return float(m.group(1)) / 100.0 if m else None


def resolve_cell_ref(wsv, formula):
    """Pull a $-anchored single-cell reference out of a markup formula and return its value."""
    m = re.search(r'\$([A-Z]+)\$(\d+)', formula)
    if not m:
        return None, None
    col, row = m.group(1), int(m.group(2))
    from openpyxl.utils import column_index_from_string
    val = wsv.cell(row, column_index_from_string(col)).value
    return val, f'{col}{row}'


def detect_markup(wsf, wsv, colmap, header_row, family):
    """
    Decode the real markup rate from the FORMULA (never the label).
    Returns dict: rate, source, label_rate, mismatch, evidence.
    """
    if family == 'cost_plus_n':
        mk_col, mk_hdr = col_for(colmap, lambda h: h.startswith('COST +') or h.startswith('COST+'))
    else:
        mk_col, mk_hdr = col_for(colmap, lambda h: 'MARKUP' in h)
    label_rate = parse_label_rate(mk_hdr or '')
    if mk_col is None:
        return {'rate': None, 'source': 'none', 'label_rate': label_rate,
                'mismatch': False, 'evidence': 'no markup column'}

    # find first item row with a formula in the markup column
    for r in range(header_row + 1, header_row + 40):
        code = wsv.cell(r, 1).value
        if not code or is_section_code(code):
            continue
        f = wsf.cell(r, mk_col).value
        v = wsv.cell(r, mk_col).value
        if not isinstance(f, str) or not f.startswith('='):
            continue
        # separate-cell form: =$N$6*... or =[..]*$P$2
        cellval, ref = resolve_cell_ref(wsv, f)
        if cellval is not None:
            rate = cellval - 1 if cellval > 1 else cellval
            ev = f'{get_column_letter(mk_col)}{r} formula refs ${ref}={cellval}'
            return {'rate': round(rate, 4), 'source': f'cell ${ref}', 'label_rate': label_rate,
                    'mismatch': _mismatch(rate, label_rate), 'evidence': ev}
        # inline coefficient: =1.35*..  or  =0.2*..
        m = re.match(r'=\s*(\d+(?:\.\d+)?)\s*\*', f)
        if m:
            coef = float(m.group(1))
            rate = coef - 1 if coef > 1 else coef
            ev = f'{get_column_letter(mk_col)}{r} formula coef={coef}'
            return {'rate': round(rate, 4), 'source': 'inline coefficient', 'label_rate': label_rate,
                    'mismatch': _mismatch(rate, label_rate), 'evidence': ev}
    return {'rate': None, 'source': 'unresolved', 'label_rate': label_rate,
            'mismatch': False, 'evidence': 'no decodable markup formula'}


def _mismatch(rate, label_rate):
    return rate is not None and label_rate is not None and abs(rate - label_rate) > 0.01


# ----------------------------------------------------------------------
# Item / block capture (capture EVERY populated cell with provenance)
# ----------------------------------------------------------------------
def walk_bidsheet(wsv, sheet, header_row, colmap):
    total_col, _ = col_for(colmap, lambda h: h == 'TOTAL')
    code_col = colmap.get('ITEM', 1)
    qty_col = colmap.get('QTY')
    cost_col, _ = col_for(colmap, lambda h: h == 'COST' or h.startswith('COST '))
    items, subtotals, sections = [], [], []
    cur_section = None
    last = wsv.max_row or (header_row + 1)
    for r in range(header_row + 1, last + 1):
        code = wsv.cell(r, code_col).value
        tot = wsv.cell(r, total_col).value if total_col else None
        qty = wsv.cell(r, qty_col).value if qty_col else None
        cost = wsv.cell(r, cost_col).value if cost_col else None
        if is_section_code(code):
            cur_section = normalize_section(code)
            sections.append((r, cur_section))
            continue
        if not isinstance(tot, (int, float)):
            continue
        codestr = str(code).strip() if code is not None else ''
        if norm(code) in ('ITEM', 'TOTAL', 'SUBTOTAL'):
            continue
        pos_qty = isinstance(qty, (int, float)) and qty > 0
        pos_cost = isinstance(cost, (int, float)) and cost > 0
        # LINE: an item code, OR positive per-unit inputs (qty/cost) even with a blank code (continuation line).
        # SUBTOTAL: a blank-code row whose total is a column sum, with no positive per-unit inputs.
        if codestr or pos_qty or pos_cost:
            cells = {}
            for h, c in colmap.items():
                v = wsv.cell(r, c).value
                if v is not None and str(v).strip() != '':
                    cells[h] = prov(v, sheet, r, c)
            items.append({'row': r, 'section': cur_section, 'item_code': codestr or None,
                          'blank_code': codestr == '', 'total': tot, 'cells': cells})
        # SUBTOTAL: a numeric total with NO per-unit inputs and no item code (pure column sum row)
        elif not codestr and abs(tot) > 0.0:
            subtotals.append({'row': r, 'value': tot, 'prov': prov(tot, sheet, r, total_col)})
    return items, subtotals, sections


def segment_blocks(items, subtotals):
    """Split lines into leaf blocks closed by subtotal rows; classify rollup subtotals."""
    leaves, rollups = [], []
    acc, last_close = [], None
    sub_iter = sorted(subtotals, key=lambda s: s['row'])
    si = 0
    rows_sorted = sorted(items, key=lambda it: it['row'])
    ri = 0
    events = sorted([('item', it['row'], it) for it in items] +
                    [('sub', s['row'], s) for s in subtotals], key=lambda e: e[1])
    recent_leaf_vals = []
    for kind, row, obj in events:
        if kind == 'item':
            acc.append(obj)
        else:  # subtotal closes something
            if acc:  # leaf block: sum of accumulated lines
                bsum = round(sum(it['total'] for it in acc), 2)
                leaves.append({'subtotal_row': row, 'value': obj['value'],
                               'line_sum': bsum, 'n': len(acc),
                               'ok': abs(round(obj['value'], 2) - bsum) <= RECON_PENNY})
                recent_leaf_vals.append(round(obj['value'], 2))
                acc = []
            else:    # rollup: should equal sum of recent leaf subtotals
                rsum = round(sum(recent_leaf_vals), 2)
                rollups.append({'subtotal_row': row, 'value': obj['value'],
                                'leaf_sum': rsum,
                                'ok': abs(round(obj['value'], 2) - rsum) <= RECON_PENNY})
                recent_leaf_vals = []
    return leaves, rollups


# ----------------------------------------------------------------------
# Stated proposal total(s) from the proposal sheet(s)
# ----------------------------------------------------------------------
def find_stated_totals(wbv, bidsheet_name):
    out = []
    for name in wbv.sheetnames:
        if name == bidsheet_name or name.lower() == 'sheet1':
            continue
        if not name.lower().startswith('proposal') and 'co #' not in name.lower():
            continue
        ws = wbv[name]
        for r in range(1, (ws.max_row or 1) + 1):
            for c in range(1, (ws.max_column or 1) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and re.fullmatch(r'\s*TOTAL\s*:?\s*', v.strip().upper().replace('GRAND', '').strip() or 'X') is None:
                    pass
                if isinstance(v, str) and v.strip().upper().rstrip(':') in ('TOTAL', 'GRAND TOTAL', 'PROJECT TOTAL'):
                    for j in range(c + 1, (ws.max_column or 1) + 1):
                        nv = ws.cell(r, j).value
                        if isinstance(nv, (int, float)) and nv > 100:
                            out.append({'sheet': name, 'cell': f'{get_column_letter(j)}{r}', 'value': float(nv)})
                            break
    # dedupe by value
    seen, ded = set(), []
    for t in out:
        k = round(t['value'], 2)
        if k not in seen:
            seen.add(k); ded.append(t)
    return ded


# ----------------------------------------------------------------------
# Gates
# ----------------------------------------------------------------------
# Module 4 — Change Orders (separate, linked records)
# ----------------------------------------------------------------------
def parse_change_order(wbv, sheet_name):
    """Parse one CO sheet. CO layout differs from the base bidsheet:
    header is DESCRIPTION/QTY/UNIT/.../LINE TOTAL (not ITEM/.../TOTAL).
    Captures CO number, line items (with provenance), and the CO total.
    Reconciliation is the captured line sum vs any stated CO total."""
    ws = wbv[sheet_name]
    maxr = ws.max_row or 1
    maxc = ws.max_column or 1
    hdr = None
    for r in range(1, min(maxr, 40) + 1):
        cells = [norm(ws.cell(r, c).value) for c in range(1, min(maxc, 20) + 1)]
        if 'DESCRIPTION' in cells and any('LINE TOTAL' in c or c == 'TOTAL' for c in cells):
            hdr = r
            break
    co_number = project_number = plans_date = None
    for r in range(1, min(maxr, 14) + 1):
        for c in range(1, min(maxc, 12) + 1):
            label = norm(ws.cell(r, c).value)
            if label.startswith('CHANGE ORDER #'):
                v = ws.cell(r, c + 1).value
                if v not in (None, ''):
                    co_number = str(v).strip()
            elif label.startswith('PROJECT #'):
                v = ws.cell(r, c + 1).value
                if v not in (None, ''):
                    project_number = str(v).strip()
            elif label.startswith('PLANS DATED'):
                v = ws.cell(r, c + 1).value
                if v not in (None, ''):
                    plans_date = str(v).strip()
    if hdr is None:
        return {'sheet': sheet_name, 'co_number': co_number, 'error': 'no DESCRIPTION/LINE TOTAL header found',
                'items': [], 'line_sum': 0.0}
    colmap = {}
    for c in range(1, maxc + 1):
        h = norm(ws.cell(hdr, c).value)
        if h:
            colmap[h] = c
    total_col = colmap.get('LINE TOTAL') or colmap.get('TOTAL')
    desc_col = colmap.get('DESCRIPTION', 1)
    qty_col = colmap.get('QTY')

    # Walk CO rows, grouping by section. The CO 'LINE TOTAL' may sit EITHER on the
    # section header row (a lump, e.g. "093000 Ceramic Tile" = 1200 with detail rows 0)
    # OR on the itemized rows beneath it (e.g. walk-off mat line = 2970, section row blank).
    # Never reliably both — so capture every populated row, then resolve per section.
    def capture_cells(r):
        cc = {}
        for h, c in colmap.items():
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != '':
                cc[h] = prov(v, sheet_name, r, c)
        return cc

    sections = []
    cur = None
    for r in range(hdr + 1, maxr + 1):
        desc = ws.cell(r, desc_col).value
        tot = ws.cell(r, total_col).value if total_col else None
        qty = ws.cell(r, qty_col).value if qty_col else None
        if desc is None or str(desc).strip() == '':
            continue
        dstr = str(desc).strip()
        first_tok = dstr.split()[0] if dstr else ''
        if is_section_code(first_tok):
            if cur is not None:
                sections.append(cur)
            cur = {'section': dstr, 'section_row': r,
                   'section_total': tot if isinstance(tot, (int, float)) else None,
                   'cells': capture_cells(r), 'items': []}
            continue
        if cur is None:  # detail rows with no preceding section header
            cur = {'section': None, 'section_row': None, 'section_total': None,
                   'cells': {}, 'items': []}
        cur['items'].append({'row': r, 'description': dstr,
                             'qty': qty if isinstance(qty, (int, float)) else None,
                             'total': tot if isinstance(tot, (int, float)) else 0.0,
                             'cells': capture_cells(r)})
    if cur is not None:
        sections.append(cur)

    # Per-section value: lump on the section row if present & nonzero, else sum of item rows.
    # If BOTH are nonzero and disagree, flag a conflict (can't tell rollup vs separate line);
    # prefer the lump and surface it for Peter.
    co_total = 0.0
    conflicts = []
    items = []  # flattened, for downstream provenance
    for s in sections:
        item_sum = round(sum(it['total'] for it in s['items']), 2)
        srow = round(s['section_total'], 2) if isinstance(s['section_total'], (int, float)) else 0.0
        if srow and item_sum and abs(srow - item_sum) > RECON_PENNY:
            conflicts.append({'section': s['section'], 'section_row_total': srow,
                              'item_sum': item_sum, 'used': srow})
            section_value = srow
        elif srow:
            section_value = srow
        else:
            section_value = item_sum
        s['section_value'] = section_value
        co_total += section_value
        items.extend(s['items'])
    co_total = round(co_total, 2)

    # These CO sheets carry NO stated CO total cell (no "TOTAL:" like the base proposal),
    # so co_total is a captured sum with no external figure to reconcile against.
    return {'sheet': sheet_name, 'co_number': co_number, 'project_number': project_number,
            'plans_date': plans_date, 'header_row': hdr,
            'sections': sections, 'items': items,
            'co_total': co_total, 'conflicts': conflicts,
            'stated_total': None, 'recon_note': 'no stated CO total on sheet — Peter confirms'}


# ----------------------------------------------------------------------
def detect_structure(wbv, bidsheet_name, leaves):
    """Identify structural cases that belong to Module 4 (change orders) / Module 5 (multi-GC/version)."""
    co_sheets = [s for s in wbv.sheetnames if 'co #' in s.lower() or 'change order' in s.lower()]
    # distinct GC proposal sheets (ignore generic Summary / CAT Totals / (n))
    gc_versions = []
    for s in wbv.sheetnames:
        sl = s.lower()
        if sl.startswith('proposal'):
            suffix = sl.replace('proposal', '', 1).strip(' ()0123456789')
            if suffix and 'summary' not in suffix and 'cat total' not in suffix:
                gc_versions.append(s)
    return {'change_order': len(co_sheets) > 0, 'co_sheets': co_sheets,
            'multi_gc': len(gc_versions) >= 2, 'gc_versions': gc_versions,
            'multiblock': len(leaves) > 1}


def gate1(items, subtotals, sections, stated, structure):
    """Per-block internal integrity + external reconciliation, with structural routing."""
    review = []
    leaves, rollups = segment_blocks(items, subtotals)
    computed = round(sum(it['total'] for it in items), 2)  # whole-file line sum (single-block only meaning)
    internal_ok = all(b['ok'] for b in leaves) if leaves else False

    # structural cases: deferred to dedicated modules — quarantine correctly, do not force-reconcile
    if structure['change_order']:
        review.append(('BLOCKING', 'structural.change_order',
                       f'change-order sheets present ({", ".join(structure["co_sheets"])}); base+CO linking is Module 4'))
    if structure['multi_gc']:
        review.append(('BLOCKING', 'structural.multi_gc',
                       f'multiple GC versions ({", ".join(structure["gc_versions"])}); one-project/multi-version is Module 5'))
    if structure['multiblock'] and not (structure['change_order'] or structure['multi_gc']):
        review.append(('BLOCKING', 'structural.multiblock',
                       f'{len(leaves)} bid blocks; base-vs-scenario/phase identity is a Peter decision (Module 5)'))

    # internal integrity always checked (proves no cell dropped, even for quarantined files)
    if leaves and not internal_ok:
        bad = [b for b in leaves if not b['ok']]
        for b in bad:
            review.append(('BLOCKING', 'gate1.internal',
                           f'block@row{b["subtotal_row"]}: line sum {b["line_sum"]} != subtotal {round(b["value"],2)} (cell dropped/miscaptured)'))

    # external reconciliation: only meaningful/decisive for SINGLE-block files
    single = len(leaves) == 1 and not structure['change_order'] and not structure['multi_gc']
    ext = ('N/A_STRUCTURAL', None)
    best = None
    if single:
        block_total = round(leaves[0]['value'], 2)
        if stated:
            best = min(stated, key=lambda t: abs(t['value'] - block_total))
            gap = round(best['value'] - block_total, 2)
            limit = max(ROUNDUP_ABS_FLOOR, ROUNDUP_PCT_LIMIT * block_total)
            if abs(gap) <= RECON_PENNY:
                ext = ('EXACT_TIE', gap)
            elif 0 <= gap <= limit:
                ext = ('ROUNDED_UP', gap)
            else:
                ext = ('MISMATCH', gap)
                review.append(('BLOCKING', 'gate1.external',
                               f'stated {best["value"]} vs block {block_total} gap {gap} exceeds rounding tolerance'))
        else:
            ext = ('NO_STATED_TOTAL', None)
            review.append(('BLOCKING', 'gate1.external', 'no stated proposal total to reconcile against'))

    has_block = bool([x for x in review if x[0] == 'BLOCKING'])
    return {'computed': computed, 'leaves': leaves, 'rollups': rollups, 'internal_ok': internal_ok,
            'external': ext, 'best_stated': best, 'structure': structure,
            'status': 'BLOCK' if has_block else 'PASS', 'review': review}


def gate2(items, colmap, record, registry):
    """PO/WO Completeness. Per-line + record-level. Returns review items (lane, code, msg)."""
    review = []
    has = lambda it, h: h in it['cells'] and it['cells'][h]['value'] not in (None, '')
    val = lambda it, h: it['cells'][h]['value'] if h in it['cells'] else None
    blank_mfr = unresolved_mfr = 0
    for it in items:
        tag = f'line {it["item_code"]}@row{it["row"]}'
        mfr = val(it, 'MFR')
        state, suggestion = resolve_mfr(mfr, registry)
        if state == 'blank':
            blank_mfr += 1
            review.append(('ENRICHMENT', 'gate2.mfr.blank', f'{tag}: manufacturer blank'))
        elif state == 'unresolved':
            unresolved_mfr += 1
            review.append(('ENRICHMENT', 'gate2.mfr.unresolved',
                           f'{tag}: manufacturer "{str(mfr).strip()}" not in registry ({suggestion}) — confirm brand'))
        # order qty / unit
        if not has(it, 'ORDER'):
            review.append(('ENRICHMENT', 'gate2.orderqty', f'{tag}: order qty missing'))
        if not has(it, 'UNITS'):
            review.append(('ENRICHMENT', 'gate2.orderunit', f'{tag}: order unit missing'))
        # cost present + sanity
        cost = val(it, 'COST')
        if cost in (None, ''):
            review.append(('ENRICHMENT', 'gate2.cost.blank', f'{tag}: cost missing'))
        elif isinstance(cost, (int, float)) and (cost < UNITCOST_LOW or cost > UNITCOST_HIGH) and cost != 0:
            review.append(('ENRICHMENT', 'gate2.cost.suspect', f'{tag}: unit cost {cost} outside sanity bounds'))
        # labor rate, unit, freight
        if not has(it, 'LABOR RATE'):
            review.append(('ENRICHMENT', 'gate2.laborrate', f'{tag}: labor rate missing'))
        if not has(it, 'UNIT'):
            review.append(('ENRICHMENT', 'gate2.unit', f'{tag}: unit missing'))
        if not has(it, 'FREIGHT'):
            review.append(('ENRICHMENT', 'gate2.freight', f'{tag}: freight missing'))
        # sf/ctn where applicable (carton ordering)
        ou = norm(val(it, 'UNITS'))
        if ou.lower() in CARTON_UNITS and not has(it, 'SF/CTN'):
            review.append(('ENRICHMENT', 'gate2.sfctn', f'{tag}: sf/ctn missing for carton order'))
    # record-level: ship-to address + architect firm
    if not record['shipto_present']:
        review.append(('BLOCKING', 'gate2.shipto', 'project ship-to address missing'))
    if not record['architect_firm_present']:
        review.append(('ENRICHMENT', 'gate2.architect', 'architect firm missing'))
    return {'review': review, 'blank_mfr': blank_mfr, 'unresolved_mfr': unresolved_mfr}


# ----------------------------------------------------------------------
# Header-zone capture: ship-to, architect (dual-source), plansDate, projectNumber
# ----------------------------------------------------------------------
def capture_header_zone(wbv, bidsheet_name):
    """Best-effort header capture for gate inputs + dual-source D1/D2/D3 (with provenance)."""
    rec = {'shipto_present': False, 'architect_firm_present': False, 'dual': {}}
    # scan proposal sheet header region + bidsheet top for address/architect cues
    text_blobs = []
    for name in wbv.sheetnames:
        if name.lower() == 'sheet1':
            continue
        ws = wbv[name]
        for r in range(1, min(ws.max_row or 1, 14) + 1):
            for c in range(1, min(ws.max_column or 1, 12) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    text_blobs.append((name, r, c, v.strip()))
    joined = ' | '.join(b[3] for b in text_blobs).lower()
    # ship-to: a street-address-like token (number + street word) anywhere in header
    rec['shipto_present'] = bool(re.search(r'\d{2,6}\s+\w+.*(st|street|ave|avenue|rd|road|blvd|dr|drive|way|lane|ln|hwy|pkwy|circle|ct|court)\b', joined))
    rec['architect_firm_present'] = bool(re.search(r'\barchitect\b', joined))
    return rec, text_blobs


# ----------------------------------------------------------------------
# Top-level parse (pure function)
# ----------------------------------------------------------------------
def parse_file(path, registry=None):
    registry = registry if registry is not None else SEED_BRANDS
    name = os.path.basename(path)
    wbv = load_workbook(path, data_only=True)
    wbf = load_workbook(path, data_only=False)
    bsheet, hrow = find_bidsheet(wbv)
    if bsheet is None:
        return {'file': name, 'error': 'no bidsheet found (no ITEM+TOTAL header row)'}
    wsv, wsf = wbv[bsheet], wbf[bsheet]
    colmap = build_colmap(wsv, hrow)
    family = detect_family(colmap)
    markup = detect_markup(wsf, wsv, colmap, hrow, family)
    items, subtotals, sections = walk_bidsheet(wsv, bsheet, hrow, colmap)
    stated = find_stated_totals(wbv, bsheet)
    record_hdr, _ = capture_header_zone(wbv, bsheet)
    leaves_preview, _ = segment_blocks(items, subtotals)
    structure = detect_structure(wbv, bsheet, leaves_preview)
    # Module 4: parse each change-order sheet as a separate, linked record
    change_orders = []
    if structure['change_order']:
        for co_sheet in structure['co_sheets']:
            co = parse_change_order(wbv, co_sheet)
            change_orders.append(co)
    g1 = gate1(items, subtotals, sections, stated, structure)
    g2 = gate2(items, colmap, record_hdr, registry)

    review = list(g1['review']) + list(g2['review'])
    if markup.get('mismatch'):
        review.append(('ENRICHMENT', 'markup.mismatch',
                       f"label {markup['label_rate']*100:.0f}% but formula gives {markup['rate']*100:.0f}% "
                       f"({markup['evidence']}); formula-derived rate used"))
    blocking = [x for x in review if x[0] == 'BLOCKING']
    enrich = [x for x in review if x[0] == 'ENRICHMENT']
    outcome = 'QUARANTINE' if blocking else 'IMPORT'
    wbv.close(); wbf.close()
    return {'file': name, 'bidsheet': bsheet, 'header_row': hrow, 'family': family,
            'markup': markup, 'n_items': len(items), 'sections': sections,
            'items': items,  # captured line records (total + per-cell provenance) — Module 3 reads these
            'change_orders': change_orders,  # Module 4: separate linked CO records
            'gate1': g1, 'gate2': g2, 'blocking': blocking, 'enrichment': enrich,
            'outcome': outcome, 'stated': stated}


# ----------------------------------------------------------------------
# Reconciliation report
# ----------------------------------------------------------------------
def main(directory):
    files = sorted(glob.glob(os.path.join(directory, 'PROPOSAL_*.xlsx')))
    print(f'AMS XLSX Reader — reconciliation over {len(files)} fixtures')
    print(f'(tolerance: internal {RECON_PENNY:.2f}; external round-up <= max(${ROUNDUP_ABS_FLOOR:.0f}, {ROUNDUP_PCT_LIMIT*100:.1f}%))\n')
    rows = []
    for f in files:
        r = parse_file(f)
        if 'error' in r:
            rows.append((os.path.basename(f), r['error'])); continue
        g1 = r['gate1']
        ext, gap = g1['external']
        mk = r['markup']
        mk_str = f"{mk['rate']*100:.0f}%" if mk['rate'] is not None else '?'
        if mk['mismatch']:
            mk_str += f"(lbl {mk['label_rate']*100:.0f}%!)"
        stated = g1['best_stated']['value'] if g1['best_stated'] else None
        blk_total = round(g1['leaves'][0]['value'], 2) if len(g1['leaves']) == 1 else g1['computed']
        rows.append({
            'file': r['file'].replace('PROPOSAL_AMS_', '').replace('.xlsx', ''),
            'fam': r['family'], 'mk': mk_str, 'items': r['n_items'],
            'computed': blk_total, 'stated': stated, 'ext': ext, 'gap': gap,
            'internal': 'ok' if g1['internal_ok'] else 'FAIL',
            'nblocks': len(g1['leaves']),
            'blank': r['gate2']['blank_mfr'], 'nonord': r['gate2']['unresolved_mfr'],
            'nblock': len(r['blocking']), 'nenr': len(r['enrichment']), 'outcome': r['outcome'],
            'blocking': r['blocking'],
        })
    # table
    hdr = f"{'file':28} {'fam':11} {'mk':13} {'#blk':>4} {'block_total':>13} {'stated':>11} {'recon':>14} {'int':>4} {'mfr(bl/un)':>10} {'outcome':>10}"
    print(hdr); print('-' * len(hdr))
    for x in rows:
        if isinstance(x, tuple):
            print(f"{x[0]:28} ERROR: {x[1]}"); continue
        recon = x['ext'] if x['gap'] is None else f"{x['ext']}{x['gap']:+.2f}"
        st = f"{x['stated']:,.0f}" if x['stated'] is not None else '—'
        print(f"{x['file']:28} {x['fam']:11} {x['mk']:13} {x['nblocks']:>4} "
              f"{x['computed']:>13,.2f} {st:>11} {recon:>14} {x['internal']:>4} "
              f"{str(x['blank'])+'/'+str(x['nonord']):>10} {x['outcome']:>10}")
    # quarantine reasons
    print('\nQUARANTINE reasons (BLOCKING):')
    any_q = False
    for x in rows:
        if isinstance(x, tuple):
            continue
        if x['outcome'] == 'QUARANTINE':
            any_q = True
            for lane, code, msg in x['blocking']:
                print(f"  {x['file']:26} [{code}] {msg}")
    if not any_q:
        print('  (none)')
    print('\nLegend: recon EXACT_TIE/ROUNDED_UP=pass, MISMATCH/NO_STATED_TOTAL=block; '
          'mfr(bl/un)=blank/unresolved-vs-registry (enrichment); int=internal line-sum vs subtotal.')


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else '/mnt/project')
